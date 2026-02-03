import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------
# Formatting
# -------------------------
# Fills and fonts
grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
blue_fill = PatternFill(start_color="A6C9EC", end_color="A6C9EC", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

white_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)
black_font = Font(color="000000", bold=True)

center_align = Alignment(horizontal="center", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))

# US Federal Holidays 2026
FEDERAL_HOLIDAYS = pd.to_datetime([
    '2026-01-01', '2026-01-19', '2026-02-16', '2026-05-25',
    '2026-06-19', '2026-07-04', '2026-09-07', '2026-10-12',
    '2026-11-11', '2026-11-26', '2026-12-25'
])

# Splicer / Pedestal / Construction lists
SPLICER_NAMES = [
    'Alexander Quishpi','Christian Cruz','Colm Coville','Devon Youmans','Douglas Masiuk',
    'Edward Healey','Elvin Corchado','Gavin Houser','Gerado Fontanez','Jancarlos Rios',
    'Jaylee Perez','Jordi Cruz','Jorge Santiago','Kenneth Baker','Kieran Healey',
    'Matthew Dziarkowski','Nathan Vazquez','Nicholas Severino','Patrick Gibaldi',
    'Paul Greene','Shane Schuler','Tamar Rascoe','Thomas Gonzales','Zachary Wood'
]

PEDESTAL_NAMES = ['Caden Kiddy','Jake Lamoureux','Ayden Lerchen']
CONSTRUCTION_NAMES = [
    'Nick Dziarkowski','Paulo Machado','Donovan Frost','Jessie Gott','Deyvis Peralta',
    'Charles Eldridge','Hector Mercado-Perez'
]



def export_weekly_report(df: pd.DataFrame, output_path: str):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if df.empty:
        print(f"No data to export for {output_path}")
        return

    if "date" not in df.columns:
        raise ValueError(f"'date' column missing. Columns present: {list(df.columns)}")

    # Ensure date is datetime
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    # Full name and category
    df['full_name'] = df['first_name'].str.title() + " " + df['last_name'].str.title()

    def assign_category(name):
        if name in SPLICER_NAMES:
            return 'Splicer'
        elif name in PEDESTAL_NAMES:
            return 'Pedestal'
        elif name in CONSTRUCTION_NAMES:
            return 'Construction'
        else:
            return 'Other'

    df['category_type'] = df['full_name'].apply(assign_category)
    df['separator'] = df['category_type'].eq('Construction')

    splicer_df = df[df['category_type'] == 'Splicer']
    pedestal_construction_df = df[df['category_type'].isin(['Pedestal', 'Construction'])]

    if not splicer_df.empty:
        _export_excel(splicer_df, output_path.with_name('Splicer_Weekly_Report.xlsx'))

    if not pedestal_construction_df.empty:
        _export_excel(pedestal_construction_df, output_path.with_name('Pedestal_Construction_Weekly_Report.xlsx'))


def _export_excel(df: pd.DataFrame, output_path: Path):
    wb = Workbook()
    ws = wb.active

    # Safe sheet title
    max_date = df['date'].max()
    week_end_date = max_date.strftime("Week Ending %m-%d-%Y") if pd.notnull(max_date) else "Week Ending Unknown"
    ws.title = week_end_date.replace('/', '-')

    # Employees
    employees = df['full_name'].drop_duplicates().tolist()
    ws.append(['Date', 'Subrow'] + employees)
    ws['B1'].font = white_font
    ws.freeze_panes = 'C2'

    subrows = ['paychex_start','samsara_start','time_difference','category',
               'paychex_end','samsara_end','time_difference','category']

    for date, date_group in df.groupby('date'):
        date_str = date.strftime("%A, %B %d, %Y") if pd.notnull(date) else "Unknown Date"
        row_dicts = {sub: {emp: '' for emp in employees} for sub in subrows}

        for _, r in date_group.iterrows():
            emp = r['full_name']
            row_dicts['paychex_start'][emp] = r.get('paychex_start')
            row_dicts['samsara_start'][emp] = r.get('start_samsara')
            row_dicts['paychex_end'][emp] = r.get('paychex_end')
            row_dicts['samsara_end'][emp] = r.get('end_samsara')

            for label, p_col, s_col in [('start','paychex_start','start_samsara'),
                                        ('end','paychex_end','end_samsara')]:
                p, s = r.get(p_col), r.get(s_col)
                if pd.notnull(p) and pd.notnull(s):
                    diff = abs((p - s).total_seconds()) / 60
                    row_dicts['time_difference'][emp] = diff
                    if diff >= 120:
                        row_dicts['category'][emp] = 'Outrageous'
                    elif diff >= 60:
                        row_dicts['category'][emp] = 'Large'
                    elif diff >= 30:
                        row_dicts['category'][emp] = 'Slight'
                    else:
                        row_dicts['category'][emp] = 'Within Reason'

        start_row = ws.max_row + 1
        for sub in subrows:
            ws.append([date_str, sub] + [row_dicts[sub][emp] for emp in employees])
        end_row = ws.max_row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(start_row,1).alignment = center_align

    # Formatting: colors, alignment, borders
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column),2):
        for cell in row:
            val = cell.value
            if val == 'Outrageous': cell.fill = red_fill
            elif val == 'Large': cell.fill = yellow_fill
            elif val == 'Slight': cell.fill = blue_fill
            if isinstance(val,pd.Timestamp):
                cell.number_format = 'h:mm AM/PM'
            cell.alignment = center_align
            cell.border = thin_border

    # Bold headers and thick borders
    for col_idx, col_cells in enumerate(ws.columns,1):
        for row_idx, cell in enumerate(col_cells,1):
            if row_idx==1 or col_idx<=2:
                cell.font = bold_font
                cell.border = thick_border

    # Hide time_difference subrows
    for r_idx, row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2,max_col=2),2):
        if row[0].value == 'time_difference':
            ws.row_dimensions[r_idx].hidden = True

    # -------------------------
    # Set fixed column widths
    # -------------------------
    ws.column_dimensions['A'].width = 44
    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 21

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Weekly report exported to {output_path}")

    wb = Workbook()
    ws = wb.active

    # Safe sheet title
    max_date = df['date'].max()
    week_end_date = max_date.strftime("Week Ending %m-%d-%Y") if pd.notnull(max_date) else "Week Ending Unknown"
    ws.title = week_end_date.replace('/', '-')

    # Employees
    employees = df['full_name'].drop_duplicates().tolist()
    ws.append(['Date', 'Subrow'] + employees)
    ws['B1'].font = white_font
    ws.freeze_panes = 'C2'

    subrows = ['paychex_start','samsara_start','time_difference','category',
               'paychex_end','samsara_end','time_difference','category']

    for date, date_group in df.groupby('date'):
        date_str = date.strftime("%A, %B %d, %Y") if pd.notnull(date) else "Unknown Date"
        row_dicts = {sub: {emp: '' for emp in employees} for sub in subrows}

        for _, r in date_group.iterrows():
            emp = r['full_name']
            row_dicts['paychex_start'][emp] = r.get('paychex_start')
            row_dicts['samsara_start'][emp] = r.get('start_samsara')
            row_dicts['paychex_end'][emp] = r.get('paychex_end')
            row_dicts['samsara_end'][emp] = r.get('end_samsara')

            for label, p_col, s_col in [('start','paychex_start','start_samsara'),
                                        ('end','paychex_end','end_samsara')]:
                p, s = r.get(p_col), r.get(s_col)
                if pd.notnull(p) and pd.notnull(s):
                    diff = abs((p - s).total_seconds()) / 60
                    row_dicts['time_difference'][emp] = diff
                    if diff >= 120:
                        row_dicts['category'][emp] = 'Outrageous'
                    elif diff >= 60:
                        row_dicts['category'][emp] = 'Large'
                    elif diff >= 30:
                        row_dicts['category'][emp] = 'Slight'
                    else:
                        row_dicts['category'][emp] = 'Within Reason'

        start_row = ws.max_row + 1
        for sub in subrows:
            ws.append([date_str, sub] + [row_dicts[sub][emp] for emp in employees])
        end_row = ws.max_row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(start_row,1).alignment = center_align

    # Formatting: colors, alignment, borders
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column),2):
        for cell in row:
            val = cell.value
            if val == 'Outrageous': cell.fill = red_fill
            elif val == 'Large': cell.fill = yellow_fill
            elif val == 'Slight': cell.fill = blue_fill
            if isinstance(val,pd.Timestamp):
                cell.number_format = 'h:mm AM/PM'
            cell.alignment = center_align
            cell.border = thin_border

    # Bold headers and thick borders
    for col_idx, col_cells in enumerate(ws.columns,1):
        for row_idx, cell in enumerate(col_cells,1):
            if row_idx==1 or col_idx<=2:
                cell.font = bold_font
                cell.border = thick_border

    # Hide time_difference subrows
    for r_idx, row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2,max_col=2),2):
        if row[0].value == 'time_difference':
            ws.row_dimensions[r_idx].hidden = True

    # -------------------------
    # Auto-adjust column widths
    # -------------------------
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # add padding

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Weekly report exported to {output_path}")

    wb = Workbook()
    ws = wb.active

    # Safe sheet title
    max_date = df['date'].max()
    week_end_date = max_date.strftime("Week Ending %m-%d-%Y") if pd.notnull(max_date) else "Week Ending Unknown"
    ws.title = week_end_date.replace('/', '-')

    # Employees
    employees = df['full_name'].drop_duplicates().tolist()
    ws.append(['Date', 'Subrow'] + employees)
    ws['B1'].font = white_font
    ws.freeze_panes = 'C2'

    subrows = ['paychex_start','samsara_start','time_difference','category',
               'paychex_end','samsara_end','time_difference','category']

    for date, date_group in df.groupby('date'):
        date_str = date.strftime("%A, %B %d, %Y") if pd.notnull(date) else "Unknown Date"
        row_dicts = {sub: {emp: '' for emp in employees} for sub in subrows}

        for _, r in date_group.iterrows():
            emp = r['full_name']
            row_dicts['paychex_start'][emp] = r.get('paychex_start')
            row_dicts['samsara_start'][emp] = r.get('start_samsara')
            row_dicts['paychex_end'][emp] = r.get('paychex_end')
            row_dicts['samsara_end'][emp] = r.get('end_samsara')

            for label, p_col, s_col in [('start','paychex_start','start_samsara'),
                                        ('end','paychex_end','end_samsara')]:
                p, s = r.get(p_col), r.get(s_col)
                if pd.notnull(p) and pd.notnull(s):
                    diff = abs((p - s).total_seconds()) / 60
                    row_dicts['time_difference'][emp] = diff
                    if diff >= 120:
                        row_dicts['category'][emp] = 'Outrageous'
                    elif diff >= 60:
                        row_dicts['category'][emp] = 'Large'
                    elif diff >= 30:
                        row_dicts['category'][emp] = 'Slight'
                    else:
                        row_dicts['category'][emp] = 'Within Reason'

        start_row = ws.max_row + 1
        for sub in subrows:
            ws.append([date_str, sub] + [row_dicts[sub][emp] for emp in employees])
        end_row = ws.max_row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(start_row,1).alignment = center_align

    # Formatting: colors, alignment, borders
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column),2):
        for cell in row:
            val = cell.value
            if val == 'Outrageous': cell.fill = red_fill
            elif val == 'Large': cell.fill = yellow_fill
            elif val == 'Slight': cell.fill = blue_fill
            if isinstance(val,pd.Timestamp):
                cell.number_format = 'h:mm AM/PM'
            cell.alignment = center_align
            cell.border = thin_border

    # Bold headers and thick borders
    for col_idx, col_cells in enumerate(ws.columns,1):
        for row_idx, cell in enumerate(col_cells,1):
            if row_idx==1 or col_idx<=2:
                cell.font = bold_font
                cell.border = thick_border

    # Hide time_difference subrows
    for r_idx, row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2,max_col=2),2):
        if row[0].value == 'time_difference':
            ws.row_dimensions[r_idx].hidden = True

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Weekly report exported to {output_path}")